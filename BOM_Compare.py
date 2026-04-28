import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------- 配置参数（根据你的文件修改） ----------------------
# 第一个BOM（Part Reference）文件路径
file1_path = 'bom1.xlsx'
# 第二个BOM（位号）文件路径
file2_path = 'bom2.xlsx'
# 输出文件路径
output_path = 'bom对比结果.xlsx'

# 列名配置（必须和你Excel里的表头一致）
part_ref_col = 'Part Reference'  # 文件1中的Part Reference列名
ref_des_col = '位号'              # 文件2中的位号列名

# ---------------------- 读取并处理数据 ----------------------
# 读取文件1的Part Reference
df1 = pd.read_excel(file1_path)
part_ref_list = df1[part_ref_col].dropna().astype(str).str.strip().tolist()
part_ref_set = set(part_ref_list)

# 读取文件2的位号，并按逗号拆分所有位号
df2 = pd.read_excel(file2_path)
ref_des_series = df2[ref_des_col].dropna().astype(str)
ref_des_list = []
for des_str in ref_des_series:
    # 按逗号分割位号，去除空值和空格
    des_list = [d.strip() for d in des_str.split(',') if d.strip()]
    ref_des_list.extend(des_list)
ref_des_set = set(ref_des_list)

# ---------------------- 对比差异 ----------------------
# 位号里有，part reference里没有的
only_in_ref_des = sorted(list(ref_des_set - part_ref_set))
# part reference里有，位号里没有的
only_in_part_ref = sorted(list(part_ref_set - ref_des_set))

# 统一长度，短的用空字符串填充
max_len = max(len(only_in_ref_des), len(only_in_part_ref))
only_in_ref_des += [''] * (max_len - len(only_in_ref_des))
only_in_part_ref += [''] * (max_len - len(only_in_part_ref))

# 构建结果DataFrame
result_df = pd.DataFrame({
    '仅位号存在的位号': only_in_ref_des,
    '仅Part Reference存在的位号': only_in_part_ref
})

# ---------------------- 写入Excel并高亮显示 ----------------------
wb = Workbook()
ws = wb.active
ws.title = "BOM对比结果"

# 写入表头和数据
for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        # 高亮显示有差异的单元格（非空值）
        if r_idx > 1 and value != '':
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 调整列宽
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 35

# 保存文件
wb.save(output_path)
print(f"对比完成！结果已保存到：{output_path}")
print(f"仅位号存在的数量：{len([x for x in only_in_ref_des if x])}")
print(f"仅Part Reference存在的数量：{len([x for x in only_in_part_ref if x])}")
